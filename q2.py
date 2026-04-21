
from __future__ import annotations
import json
import random
import math
from pathlib import Path
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List, Dict, Tuple


# Excel文件路径
EXCEL_PATH = Path("参考算例.xlsx")
# 客户列表（1-15）
CUSTOMERS = list(range(1, 16))
# 随机种子（保证结果可复现）
SEED = 20260419
# 惩罚系数
EARLY_PENALTY = 10    # 早到惩罚
LATE_PENALTY = 20     # 晚到惩罚


@dataclass
class Node:
    """节点属性类"""
    lower_bound: int      # 时间窗下限
    upper_bound: int      # 时间窗上限
    service_time: int     # 服务时长
    demand: int           # 需求量


@dataclass
class Instance:
    """问题实例类"""
    travel_time_matrix: List[List[int]]  # 旅行时间矩阵
    node_identifiers: List[int]          # 节点ID列表
    node_dict: Dict[int, Node]           # 节点ID到节点属性的映射


def load_instance() -> Instance:

    # 加载Excel工作簿
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    worksheet_nodes = workbook["节点属性信息"]
    worksheet_time = workbook["旅行时间矩阵"]

    # 读取节点属性，构建节点字典
    node_dict = {}
    for row in worksheet_nodes.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        node_id = int(row[0])
        node_dict[node_id] = Node(
            lower_bound=int(row[1]),
            upper_bound=int(row[2]),
            service_time=int(row[3]),
            demand=int(row[4]) if row[4] else 0
        )

    # 读取旅行时间矩阵的节点ID列表
    node_identifiers = [
        int(worksheet_time.cell(1, column).value)
        for column in range(2, worksheet_time.max_column + 1)
    ]

    # 读取旅行时间矩阵数据
    travel_time_matrix = [
        [
            int(worksheet_time.cell(row, column).value)
            for column in range(2, worksheet_time.max_column + 1)
        ]
        for row in range(2, worksheet_time.max_row + 1)
    ]

    return Instance(
        travel_time_matrix=travel_time_matrix,
        node_identifiers=node_identifiers,
        node_dict=node_dict
    )


def get_travel_time(from_node: int, to_node: int, instance: Instance) -> int:

    from_index = instance.node_identifiers.index(from_node)
    to_index = instance.node_identifiers.index(to_node)
    return instance.travel_time_matrix[from_index][to_index]


def route_evaluate(route: List[int], instance: Instance) -> dict:
    # 构建完整路线（首尾加上起点0）
    full_route = [0] + route + [0]

    current_time = 0           # 当前时间（累计旅行时间+服务时间）
    total_travel_time = 0     # 总运输时间
    total_penalty = 0         # 总惩罚值
    customer_details = []      # 每客户的详细数据

    # 遍历路线的每一段
    for prev_node, node in zip(full_route[:-1], full_route[1:], strict=True):
        # 获取该段旅行时间并累加到总运输时间
        leg_time = get_travel_time(prev_node, node, instance)
        total_travel_time += leg_time

        # 跳过配送中心（不需要计算时间窗惩罚）
        if node == 0:
            continue

        # 更新当前时间（到达该节点）
        current_time += leg_time
        service_start_time = int(current_time)

        # 获取该节点的时间窗信息
        node_info = instance.node_dict[node]
        lower_bound = node_info.lower_bound
        upper_bound = node_info.upper_bound

        # 计算时间窗违反量
        early_violation = max(lower_bound - service_start_time, 0)   # 早到违反时间
        late_violation = max(service_start_time - upper_bound, 0)     # 晚到违反时间

        # 计算惩罚值（二次惩罚）
        penalty_value = EARLY_PENALTY * early_violation ** 2 + LATE_PENALTY * late_violation ** 2

        # 累加总惩罚
        total_penalty += penalty_value

        # 记录该客户的详细数据
        customer_details.append({
            "node": node,
            "arrival": int(current_time),           # 到达时间
            "start": service_start_time,            # 开始服务时间
            "lb": lower_bound,                      # 时间窗下限
            "ub": upper_bound,                      # 时间窗上限
            "early": early_violation,               # 早到违反量
            "late": late_violation,                 # 晚到违反量
            "penalty": penalty_value                # 该客户惩罚值
        })

        # 加上服务时间
        current_time += node_info.service_time

    # 返回评估结果
    return {
        "route": full_route,
        "travel_time": total_travel_time,
        "time_window_penalty": total_penalty,
        "objective": total_travel_time + total_penalty,
        "per_customer": customer_details
    }


def generate_2opt_neighbors(route: List[int]) -> List[List[int]]:
    route_length = len(route)
    candidates = []
    seen_routes = set()  # 用于去重

    # 遍历所有(i,j)组合
    for i in range(route_length):
        for j in range(i + 1, route_length):
            # 操作1：交换i和j位置的节点
            candidate_swap = route[:]
            candidate_swap[i], candidate_swap[j] = candidate_swap[j], candidate_swap[i]
            if tuple(candidate_swap) not in seen_routes:
                seen_routes.add(tuple(candidate_swap))
                candidates.append(candidate_swap)

            # 操作2：反转i到j之间的节点
            candidate_reverse = route[:]
            candidate_reverse[i:j + 1] = list(reversed(candidate_reverse[i:j + 1]))
            if tuple(candidate_reverse) not in seen_routes:
                seen_routes.add(tuple(candidate_reverse))
                candidates.append(candidate_reverse)

    return candidates


def local_search(route: List[int], instance: Instance, max_iterations: int = 50) -> Tuple[List[int], dict]:
    best_route = route[:]
    best_value = route_evaluate(best_route, instance)["objective"]

    # 主循环
    for iteration in range(max_iterations):
        improved = False

        # 遍历所有邻域解
        for candidate in generate_2opt_neighbors(best_route):
            candidate_value = route_evaluate(candidate, instance)["objective"]
            if candidate_value + 1e-9 < best_value:
                best_route = candidate
                best_value = candidate_value
                improved = True

        # 如果没有改进，退出
        if not improved:
            break

    return best_route, route_evaluate(best_route, instance)


def insert_greedy(customer_order: List[int], instance: Instance) -> List[int]:
    route = []

    # 依次处理每个客户
    for customer in customer_order:
        if not route:
            # 第一个客户直接作为起点
            route = [customer]
            continue

        best_candidate = None
        best_key = None

        # 尝试所有可能的插入位置
        for position in range(len(route) + 1):
            candidate = route[:]
            candidate.insert(position, customer)
            evaluation = route_evaluate(candidate, instance)
            key = (evaluation["objective"], evaluation["travel_time"])

            if best_key is None or key < best_key:
                best_key = key
                best_candidate = candidate

        route = best_candidate

    return route


def nearest_neighbor(start_node: int, instance: Instance) -> List[int]:
    remaining_customers = set(CUSTOMERS)
    current_node = start_node
    route = []

    # 循环直到所有客户都被访问
    while remaining_customers:
        # 选择距离最近的客户
        next_node = min(
            remaining_customers,
            key=lambda x: get_travel_time(current_node, x, instance)
        )
        route.append(next_node)
        remaining_customers.remove(next_node)
        current_node = next_node

    return route


def simulated_annealing(initial_route: List[int], instance: Instance, random_seed: int) -> Tuple[List[int], dict]:
    random_generator = random.Random(random_seed)

    best_route = initial_route[:]
    best_cost = route_evaluate(best_route, instance)["objective"]

    current_route = best_route[:]
    current_cost = best_cost

    temperature = 500.0  # 初始温度

    # 降温过程
    while temperature > 0.5:
        # 每个温度下的迭代次数
        for _ in range(100):
            # 随机选择两个位置进行交换
            i = random_generator.randrange(len(current_route))
            j = random_generator.randrange(len(current_route))

            candidate = current_route[:]
            candidate[i], candidate[j] = candidate[j], candidate[i]

            candidate_cost = route_evaluate(candidate, instance)["objective"]
            delta = candidate_cost - current_cost

            # 如果是更优解，接受
            if delta < 0:
                current_route = candidate
                current_cost = candidate_cost
                if current_cost < best_cost:
                    best_route = current_route[:]
                    best_cost = current_cost
            # 如果是劣解，按概率接受
            elif random_generator.random() < math.exp(-delta / temperature):
                current_route = candidate
                current_cost = candidate_cost

        # 降温
        temperature *= 0.97

    return best_route, route_evaluate(best_route, instance)


def multi_start_solve(instance: Instance, num_solutions: int = 10) -> List[dict]:

    random_generator = random.Random(SEED)
    all_solutions = []
    seen_keys = set()  # 用于去重

    def add_solution(route: List[int]):
        """添加解（去重）"""
        evaluation = route_evaluate(route, instance)
        key = (evaluation["objective"], tuple(route))
        if key not in seen_keys:
            seen_keys.add(key)
            all_solutions.append(evaluation)

    # ========== 策略1：多种排序贪心插入 + 2-opt ==========
    print("  [1/3] 多种排序贪心插入...")

    # 生成5种不同的客户排序方式
    customer_orders = [
        # 按时间窗上界排序
        sorted(CUSTOMERS, key=lambda i: (instance.node_dict[i].upper_bound, instance.node_dict[i].lower_bound, i)),
        # 按时间窗下界排序
        sorted(CUSTOMERS, key=lambda i: (instance.node_dict[i].lower_bound, instance.node_dict[i].upper_bound, i)),
        # 按时间窗宽度排序
        sorted(CUSTOMERS, key=lambda i: (instance.node_dict[i].upper_bound - instance.node_dict[i].lower_bound, i)),
        # 按时间窗中心排序
        sorted(CUSTOMERS, key=lambda i: (instance.node_dict[i].upper_bound + instance.node_dict[i].lower_bound, i)),
        # 按到配送中心距离排序
        sorted(CUSTOMERS, key=lambda i: (get_travel_time(0, i, instance), instance.node_dict[i].upper_bound, i)),
    ]

    for order in customer_orders:
        route = insert_greedy(order, instance)
        refined_route, _ = local_search(route, instance, max_iterations=50)
        add_solution(refined_route)

    # ========== 策略2：最近邻构造 + 2-opt ==========
    print("  [2/3] 最近邻构造...")

    for start_customer in CUSTOMERS:
        route = nearest_neighbor(start_customer, instance)
        refined_route, _ = local_search(route, instance, max_iterations=50)
        add_solution(refined_route)

    # ========== 策略3：模拟退火 + 2-opt ==========
    print("  [3/3] 模拟退火重启...")

    for sa_seed in range(SEED, SEED + 20):
        initial_route = CUSTOMERS[:]
        random_generator.shuffle(initial_route)
        sa_route, _ = simulated_annealing(initial_route, instance, sa_seed)
        refined_route, _ = local_search(sa_route, instance, max_iterations=50)
        add_solution(refined_route)

    # 按目标值排序
    all_solutions.sort(key=lambda x: (x["objective"], x["travel_time"], x["route"]))
    return all_solutions[:num_solutions]


def print_solution(solution_index: int, solution: dict) -> None:

    print(f"\n{'─' * 70}")
    print(f"解 #{solution_index}  |  obj={solution['objective']}  |  travel={solution['travel_time']}  |  penalty={solution['time_window_penalty']}")
    print(f"  路线: {solution['route']}")

    # 打印每个客户的详细数据
    for customer in solution["per_customer"]:
        flag = ""
        if customer["early"] > 0:
            flag = f" ←早违{customer['early']}×10={10*customer['early']**2}"
        elif customer["late"] > 0:
            flag = f" ←晚违{customer['late']}×20={20*customer['late']**2}"
        print(f"    客户{customer['node']:2d} | 到达{customer['arrival']:3d} | 开始{customer['start']:3d} | "
              f"窗[{customer['lb']:2d},{customer['ub']:2d}]{flag}")


def main():
    print("=" * 70)
    print("问题2：带时间窗惩罚的单车辆调度（前15个客户）")
    print("本地启发式 → 生成10个较优解（后续量子机选最优）")
    print("=" * 70)

    # 加载实例数据
    instance = load_instance()
    print(f"\n数据加载完成: 配送中心=0, 客户1-15, 时间窗惩罚: E={EARLY_PENALTY}, L={LATE_PENALTY}")

    print("\n开始求解...")
    solutions = multi_start_solve(instance, num_solutions=10)

    print(f"\n求解完成！共 {len(solutions)} 个解")
    print("=" * 70)

    for index, solution in enumerate(solutions, 1):
        print_solution(index, solution)

    # 构建结果字典
    result = {
        "problem": 2,
        "scope": {
            "depot": 0,
            "customer_ids": CUSTOMERS,
            "customer_count": len(CUSTOMERS),
            "consider_time_window_penalty": True,
            "consider_capacity": False
        },
        "strategy": {
            "name": "多策略本地启发式（贪心+最近邻+SA + 2-opt局部搜索）",
            "description": "前15客户，生成10个较优解供后续量子机选优",
            "methods": ["5种排序贪心插入", "15个起始点最近邻", "20次SA重启+2-opt"],
        },
        "top_solutions": solutions,
    }

    # 保存结果到JSON文件
    output_path = Path(r"output_json/q2_local_solutions.json")
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n结果已保存: {output_path}")

    # 打印汇总表
    print("\n" + "=" * 70)
    print("Top10汇总:")
    print(f"{'#':>3} {'objective':>12} {'travel_time':>12} {'penalty':>12}")
    print("-" * 52)

    for index, solution in enumerate(solutions, 1):
        print(f"{index:>3} {solution['objective']:>12} {solution['travel_time']:>12} {solution['time_window_penalty']:>12}")

    print("\n参考最优: travel_time=31, penalty=84090, objective=84121")
    print("  路线: [0, 2, 13, 6, 5, 8, 7, 11, 10, 1, 9, 3, 12, 4, 15, 14, 0]")


if __name__ == "__main__":
    main()
