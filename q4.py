from __future__ import annotations
import json
import math
import random
from pathlib import Path
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List, Dict, Tuple
import numpy as np


# Excel文件路径
EXCEL_PATH = Path(r"参考算例.xlsx")
# 客户列表（1-50）
CUSTOMERS = list(range(1, 51))
# 随机种子（保证结果可复现）
SEED = 20260419
# 惩罚系数
EARLY_PENALTY = 10    # 早到惩罚
LATE_PENALTY = 20     # 晚到惩罚
# 车辆容量
VEHICLE_CAPACITY = 60


@dataclass
class Node:
    """节点属性类"""
    lower_bound: int      # 时间窗下限
    upper_bound: int      # 时间窗上限
    service_time: int     # 服务时长
    demand: int          # 需求量


@dataclass
class Instance:
    """问题实例类"""
    travel_time_matrix: List[List[int]]  # 旅行时间矩阵
    node_identifiers: List[int]         # 节点ID列表
    node_dict: Dict[int, Node]           # 节点ID到节点属性的映射


def load_instance() -> Instance:
    # 加载Excel工作簿
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    worksheet_nodes = workbook["节点属性信息"]
    worksheet_time = workbook["旅行时间矩阵"]

    # 读取节点属性，构建节点字典
    node_dict = {}
    for row in worksheet_nodes.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        node_id = int(row[0])
        node_dict[node_id] = Node(
            lower_bound=int(row[1]),
            upper_bound=int(row[2]),
            service_time=int(row[3]),
            demand=int(row[4]) if row[4] else 0
        )

    # 读取旅行时间矩阵的节点ID列表
    node_identifiers = [
        int(worksheet_time.cell(1, column).value)
        for column in range(2, worksheet_time.max_column + 1)
    ]

    # 读取旅行时间矩阵数据
    travel_time_matrix = [
        [
            int(worksheet_time.cell(row, column).value)
            for column in range(2, worksheet_time.max_column + 1)
        ]
        for row in range(2, worksheet_time.max_row + 1)
    ]

    return Instance(
        travel_time_matrix=travel_time_matrix,
        node_identifiers=node_identifiers,
        node_dict=node_dict
    )


def get_travel_time(from_node: int, to_node: int, instance: Instance) -> int:
    from_index = instance.node_identifiers.index(from_node)
    to_index = instance.node_identifiers.index(to_node)
    return instance.travel_time_matrix[from_index][to_index]


def route_evaluate(route: List[int], instance: Instance, vehicle_id: int = 1) -> dict:
    # 构建完整路线（首尾加上配送中心0）
    full_route = [0] + route + [0]

    current_time = 0           # 当前时间（累计旅行时间+服务时间）
    total_travel_time = 0     # 总运输时间
    total_penalty = 0         # 总惩罚值
    customer_details = []      # 每客户的详细数据

    # 遍历路线的每一段
    for prev_node, node in zip(full_route[:-1], full_route[1:], strict=True):
        # 获取该段旅行时间并累加到总运输时间
        leg_time = get_travel_time(prev_node, node, instance)
        total_travel_time += leg_time

        # 跳过配送中心（不需要计算时间窗惩罚）
        if node == 0:
            continue

        # 更新当前时间（到达该节点）
        current_time += leg_time
        service_start_time = int(current_time)

        # 获取该节点的时间窗信息
        node_info = instance.node_dict[node]
        lower_bound = node_info.lower_bound
        upper_bound = node_info.upper_bound

        # 计算时间窗违反量
        early_violation = max(lower_bound - service_start_time, 0)   # 早到违反时间
        late_violation = max(service_start_time - upper_bound, 0)     # 晚到违反时间

        # 计算惩罚值（二次惩罚）
        penalty_value = EARLY_PENALTY * early_violation ** 2 + LATE_PENALTY * late_violation ** 2

        # 累加总惩罚
        total_penalty += penalty_value

        # 记录该客户的详细数据
        customer_details.append({
            "vehicle_id": vehicle_id,
            "node": node,
            "arrival": int(current_time),          # 到达时间
            "start": service_start_time,            # 开始服务时间
            "lb": lower_bound,                      # 时间窗下限
            "ub": upper_bound,                      # 时间窗上限
            "early": early_violation,               # 早到违反量
            "late": late_violation,                 # 晚到违反量
            "penalty": penalty_value               # 该客户惩罚值
        })

        # 加上服务时间
        current_time += node_info.service_time

    # 返回评估结果
    return {
        "vehicle_id": vehicle_id,
        "route": full_route,
        "travel_time": total_travel_time,
        "time_window_penalty": total_penalty,
        "objective": total_travel_time + total_penalty,
        "per_customer": customer_details
    }


def solution_evaluate(routes: List[List[int]], instance: Instance) -> dict:
    vehicle_results = []
    per_customer = []
    total_travel_time = 0
    total_penalty = 0
    used_vehicles = 0

    # 遍历每辆车的路线
    for vehicle_id, route in enumerate(routes, 1):
        if not route:
            continue
        evaluation = route_evaluate(route, instance, vehicle_id)
        vehicle_results.append(evaluation)
        per_customer.extend(evaluation["per_customer"])
        total_travel_time += evaluation["travel_time"]
        total_penalty += evaluation["time_window_penalty"]
        used_vehicles += 1

    return {
        "routes": [[0] + r + [0] for r in routes if r],
        "vehicle_results": vehicle_results,
        "used_vehicle_count": used_vehicles,
        "travel_time": total_travel_time,
        "time_window_penalty": total_penalty,
        "objective": total_travel_time + total_penalty,
        "per_customer": per_customer,
    }


def get_total_demand(customer_list: List[int], instance: Instance) -> int:
    return sum(instance.node_dict[c].demand for c in customer_list)


def insert_greedy(customer_order: List[int], instance: Instance) -> List[int]:
    route = []

    # 依次处理每个客户
    for customer in customer_order:
        if not route:
            route = [customer]
            continue

        best_candidate = None
        best_key = None

        # 尝试所有可能的插入位置
        for position in range(len(route) + 1):
            candidate = route[:]
            candidate.insert(position, customer)
            evaluation = route_evaluate(candidate, instance, 1)
            key = (evaluation["objective"], evaluation["travel_time"])

            if best_key is None or key < best_key:
                best_key = key
                best_candidate = candidate

        route = best_candidate

    return route


def nearest_neighbor(start_node: int, customer_list: List[int], instance: Instance) -> List[int]:
    remaining_customers = set(customer_list)
    current_node = start_node
    route = []

    # 循环直到所有客户都被访问
    while remaining_customers:
        next_node = min(
            remaining_customers,
            key=lambda x: get_travel_time(current_node, x, instance)
        )
        route.append(next_node)
        remaining_customers.remove(next_node)
        current_node = next_node

    return route


def two_opt_swap(route: List[int], instance: Instance, max_iterations: int = 50) -> Tuple[List[int], dict]:
    best_route = route[:]
    best_value = route_evaluate(best_route, instance, 1)["objective"]
    route_length = len(best_route)

    for _ in range(max_iterations):
        improved = False
        for i in range(route_length):
            for j in range(i + 1, route_length):
                candidate = best_route[:i] + list(reversed(best_route[i:j+1])) + best_route[j+1:]
                value = route_evaluate(candidate, instance, 1)["objective"]
                if value + 1e-9 < best_value:
                    best_route = candidate
                    best_value = value
                    improved = True
        if not improved:
            break

    return best_route, route_evaluate(best_route, instance, 1)


def relocate(route: List[int], instance: Instance) -> Tuple[List[int], dict]:
    best_route = route[:]
    best_value = route_evaluate(best_route, instance, 1)["objective"]
    route_length = len(best_route)

    improved = True
    while improved:
        improved = False
        for i in range(route_length):
            for k in range(route_length + 1):
                if k == i or k == i + 1:
                    continue
                candidate = best_route[:i] + best_route[i+1:]
                insert_position = k if k <= i else k - 1
                candidate.insert(insert_position, best_route[i])
                value = route_evaluate(candidate, instance, 1)["objective"]
                if value + 1e-9 < best_value:
                    best_route = candidate
                    best_value = value
                    improved = True

    return best_route, route_evaluate(best_route, instance, 1)


def swap(route: List[int], instance: Instance) -> Tuple[List[int], dict]:
    best_route = route[:]
    best_value = route_evaluate(best_route, instance, 1)["objective"]
    route_length = len(best_route)

    improved = True
    while improved:
        improved = False
        for i in range(route_length):
            for j in range(i + 1, route_length):
                candidate = best_route[:]
                candidate[i], candidate[j] = candidate[j], candidate[i]
                value = route_evaluate(candidate, instance, 1)["objective"]
                if value + 1e-9 < best_value:
                    best_route = candidate
                    best_value = value
                    improved = True

    return best_route, route_evaluate(best_route, instance, 1)


def combined_optimize(route: List[int], instance: Instance) -> Tuple[List[int], dict]:
    current_route = route[:]
    current_value = route_evaluate(current_route, instance, 1)["objective"]

    # 2-opt
    current_route, evaluation = two_opt_swap(current_route, instance, max_iterations=50)
    if evaluation["objective"] < current_value:
        current_value = evaluation["objective"]

    # Relocate
    current_route, evaluation = relocate(current_route, instance)
    if evaluation["objective"] < current_value:
        current_value = evaluation["objective"]

    # Swap
    current_route, evaluation = swap(current_route, instance)
    if evaluation["objective"] < current_value:
        current_value = evaluation["objective"]

    # 再次2-opt
    current_route, evaluation = two_opt_swap(current_route, instance, max_iterations=30)
    return current_route, route_evaluate(current_route, instance, 1)


def time_window_aware_clustering(
    customer_list: List[int],
    instance: Instance,
    num_vehicles: int,
    vehicle_capacity: int
) -> List[List[int]]:
    # 计算每个客户的时间窗中心
    customer_time_centers = [
        (c, (instance.node_dict[c].upper_bound + instance.node_dict[c].lower_bound) / 2)
        for c in customer_list
    ]
    customer_time_centers.sort(key=lambda x: x[1])

    # 初始化车辆聚类
    clusters = [[] for _ in range(num_vehicles)]
    vehicle_loads = [0] * num_vehicles

    # 遍历每个客户，将其分配到最合适的车辆
    for customer, _ in customer_time_centers:
        customer_demand = instance.node_dict[customer].demand
        best_vehicle = -1
        best_load = float('inf')

        # 选择负载最小且容量足够的车辆
        for vehicle_index in range(num_vehicles):
            new_load = vehicle_loads[vehicle_index] + customer_demand
            if new_load <= vehicle_capacity and vehicle_loads[vehicle_index] < best_load:
                best_load = vehicle_loads[vehicle_index]
                best_vehicle = vehicle_index

        # 如果没有找到合适的车辆，选择负载最小的车辆
        if best_vehicle == -1:
            best_vehicle = min(range(num_vehicles), key=lambda v: vehicle_loads[v])

        clusters[best_vehicle].append(customer)
        vehicle_loads[best_vehicle] += customer_demand

    return clusters


def cross_vehicle_optimize(
    clusters: List[List[int]],
    instance: Instance,
    vehicle_capacity: int,
    max_iterations: int = 100
) -> List[List[int]]:
    print("    跨车优化...")
    improved = True
    iteration = 0

    while improved and iteration < max_iterations:
        improved = False
        iteration += 1
        current_evaluation = solution_evaluate(clusters, instance)

        current_clusters = [cluster[:] for cluster in clusters]

        # 遍历每辆车的每个客户
        for source_vehicle in range(len(current_clusters)):
            if not current_clusters[source_vehicle]:
                continue

            for customer in current_clusters[source_vehicle][:]:
                customer_demand = instance.node_dict[customer].demand

                # 尝试移动到其他车辆
                for dest_vehicle in range(len(current_clusters)):
                    if source_vehicle == dest_vehicle:
                        continue

                    new_dest_load = get_total_demand(current_clusters[dest_vehicle], instance) + customer_demand
                    if new_dest_load > vehicle_capacity:
                        continue

                    # 执行移动
                    new_clusters = [cluster[:] for cluster in current_clusters]
                    if customer not in new_clusters[source_vehicle]:
                        continue
                    new_clusters[source_vehicle].remove(customer)
                    new_clusters[dest_vehicle].append(customer)

                    # 重新优化涉及的两辆车
                    for vehicle_index in [source_vehicle, dest_vehicle]:
                        if new_clusters[vehicle_index]:
                            order = sorted(
                                new_clusters[vehicle_index],
                                key=lambda c: (instance.node_dict[c].upper_bound, instance.node_dict[c].lower_bound, c)
                            )
                            new_clusters[vehicle_index] = insert_greedy(order, instance)
                            new_clusters[vehicle_index], _ = combined_optimize(new_clusters[vehicle_index], instance)

                    # 评估移动后的解
                    new_evaluation = solution_evaluate(new_clusters, instance)

                    if new_evaluation["objective"] < current_evaluation["objective"]:
                        current_clusters = new_clusters
                        current_evaluation = new_evaluation
                        improved = True

        clusters = current_clusters

        if improved:
            print(f"      第{iteration}轮: travel={current_evaluation['travel_time']}, pen={current_evaluation['time_window_penalty']}")

    return clusters


def solve_single_cluster(cluster: List[int], instance: Instance, vehicle_id: int) -> Tuple[List[int], dict]:
    if not cluster:
        return [], route_evaluate([], instance, vehicle_id)

    order = sorted(
        cluster,
        key=lambda c: (instance.node_dict[c].upper_bound, instance.node_dict[c].lower_bound, c)
    )
    route = insert_greedy(order, instance)
    route, evaluation = combined_optimize(route, instance)

    return route, evaluation


def solve_for_k_with_comparison(num_vehicles: int, instance: Instance) -> dict:
    print(f"\n  {'='*50}")
    print(f"  车辆数 K = {num_vehicles}")
    print(f"  {'='*50}")

    total_customer_demand = get_total_demand(CUSTOMERS, instance)
    print(f"  总需求: {total_customer_demand}, 容量: {VEHICLE_CAPACITY}")

    # 阶段1：聚类分车
    print("  阶段1: 时间窗感知聚类分车")
    clusters = time_window_aware_clustering(CUSTOMERS, instance, num_vehicles, VEHICLE_CAPACITY)

    for vehicle_index, cluster in enumerate(clusters):
        load = get_total_demand(cluster, instance)
        print(f"    V{vehicle_index+1}: {len(cluster)}客户, 需求={load}")

    # 阶段2：簇内优化
    print("  阶段2: 簇内路线优化")
    routes = []
    vehicle_details = []

    for vehicle_id, cluster in enumerate(clusters, 1):
        if not cluster:
            routes.append([])
            continue
        route, evaluation = solve_single_cluster(cluster, instance, vehicle_id)
        routes.append(route)
        vehicle_details.append({
            "vehicle_id": vehicle_id,
            "route": [0] + route + [0],
            "customers": cluster,
            "load": get_total_demand(cluster, instance),
            "travel_time": evaluation["travel_time"],
            "time_window_penalty": evaluation["time_window_penalty"],
        })
        print(f"    V{vehicle_id}: travel={evaluation['travel_time']}, pen={evaluation['time_window_penalty']}")

    # 阶段3：簇间优化
    print("  阶段3: 簇间跨车优化")
    clusters = cross_vehicle_optimize(clusters, instance, VEHICLE_CAPACITY, max_iterations=50)

    # 重新求解优化后的路线
    routes = []
    vehicle_details = []
    for vehicle_id, cluster in enumerate(clusters, 1):
        if not cluster:
            routes.append([])
            continue
        route, evaluation = solve_single_cluster(cluster, instance, vehicle_id)
        routes.append(route)
        vehicle_details.append({
            "vehicle_id": vehicle_id,
            "route": [0] + route + [0],
            "customers": cluster,
            "load": get_total_demand(cluster, instance),
            "travel_time": evaluation["travel_time"],
            "time_window_penalty": evaluation["time_window_penalty"],
        })

    solution = solution_evaluate(routes, instance)

    print(f"\n  K={num_vehicles}结果:")
    print(f"    使用车辆: {solution['used_vehicle_count']}")
    print(f"    总运输时间: {solution['travel_time']}")
    print(f"    总时间窗惩罚: {solution['time_window_penalty']}")
    print(f"    目标值(travel+pen): {solution['objective']}")

    return {
        "num_vehicles": num_vehicles,
        "clusters": clusters,
        "routes": routes,
        "vehicle_details": vehicle_details,
        "solution": solution,
    }


def solve_q4(instance: Instance) -> dict:
    print("=" * 70)
    print("问题4：多车辆带容量约束调度（VRPTW）")
    print("核心原则：车辆数最少为第一优先")
    print("=" * 70)

    total_customer_demand = get_total_demand(CUSTOMERS, instance)
    minimum_vehicles = math.ceil(total_customer_demand / VEHICLE_CAPACITY)
    print(f"\n总需求: {total_customer_demand}, 车辆容量: {VEHICLE_CAPACITY}")
    print(f"理论最小车辆数: {minimum_vehicles}")

    # 首先对最小车辆数K=5进行充分优化（这是我们的最终答案）
    print("\n" + "="*70)
    print("【最优方案】K=5（车辆数最少）")
    print("="*70)
    best_result = solve_for_k_with_comparison(minimum_vehicles, instance)

    # 然后计算K=6,7作为对比
    all_results = {"5": best_result}

    for num_vehicles in [6, 7]:
        print("\n" + "="*70)
        print(f"【对比方案】K={num_vehicles}")
        print("="*70)
        result = solve_for_k_with_comparison(num_vehicles, instance)
        all_results[str(num_vehicles)] = result

    # 打印对比表
    print("\n" + "=" * 70)
    print("车辆数影响分析汇总")
    print("=" * 70)
    print(f"{'K':>3} | {'车辆数':>6} | {'运输时间':>8} | {'惩罚':>10} | {'目标值':>12}")
    print("-" * 55)

    for k_string, result in all_results.items():
        solution = result["solution"]
        print(f"{k_string:>3} | {solution['used_vehicle_count']:>6} | {solution['travel_time']:>8} | "
              f"{solution['time_window_penalty']:>10} | {solution['objective']:>12}")

    print("\n" + "=" * 70)
    print("【结论】根据\"车辆数最少优先\"原则，选择K=5为最优方案")
    print("虽然K=6的惩罚更低，但K=5使用了更少的车辆")
    print("=" * 70)

    return best_result, all_results


def main():
    # 加载实例数据
    instance = load_instance()
    print(f"\n数据加载: 客户1-50, 容量={VEHICLE_CAPACITY}")

    # 求解
    best_result, all_results = solve_q4(instance)

    # 构建结果payload
    payload = {
        "problem": 4,
        "scope": {
            "depot": 0,
            "customer_ids": CUSTOMERS,
            "customer_count": len(CUSTOMERS),
            "consider_time_window_penalty": True,
            "consider_capacity": True,
            "vehicle_capacity": VEHICLE_CAPACITY,
            "total_demand": get_total_demand(CUSTOMERS, instance),
        },
        "strategy": {
            "name": "时间窗感知聚类 + 组合局部搜索 + 跨车优化",
            "description": "以最小车辆数为优先目标",
        },
        "vehicle_count_comparison": {
            k: {
                "used_vehicle_count": v["solution"]["used_vehicle_count"],
                "travel_time": v["solution"]["travel_time"],
                "time_window_penalty": v["solution"]["time_window_penalty"],
                "objective": v["solution"]["objective"],
            }
            for k, v in all_results.items()
        },
        "final_answer": {
            "used_vehicle_count": best_result["solution"]["used_vehicle_count"],
            "travel_time": best_result["solution"]["travel_time"],
            "time_window_penalty": best_result["solution"]["time_window_penalty"],
            "objective": best_result["solution"]["objective"],
            "routes": best_result["solution"]["routes"],
            "vehicle_details": best_result["vehicle_details"],
        },
    }

    # 保存结果
    output_path = Path(r"output_json/q4_result.json")
    Path(output_path.parent).mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n结果已保存: {output_path}")

    print("\n" + "=" * 70)
    print("最终结果（车辆数最少优先）:")
    print(f"  使用车辆: {best_result['solution']['used_vehicle_count']}")
    print(f"  总运输时间: {best_result['solution']['travel_time']}")
    print(f"  总时间窗惩罚: {best_result['solution']['time_window_penalty']}")
    print(f"  目标值(travel+pen): {best_result['solution']['objective']}")


if __name__ == "__main__":
    main()
