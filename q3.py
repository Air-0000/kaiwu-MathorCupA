from __future__ import annotations
import json
import random
from pathlib import Path
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List, Dict, Tuple
from itertools import product


# Excel文件路径
EXCEL_PATH = Path("参考算例.xlsx")
# 客户列表（1-50）
CUSTOMERS = list(range(1, 51))
# 随机种子（保证结果可复现）
SEED = 20260419
# 惩罚系数
EARLY_PENALTY = 10    # 早到惩罚
LATE_PENALTY = 20     # 晚到惩罚


@dataclass
class Node:
    """节点属性类"""
    lower_bound: int      # 时间窗下限
    upper_bound: int      # 时间窗上限
    service_time: int     # 服务时长
    demand: int          # 需求量


@dataclass
class Instance:
    """问题实例类"""
    travel_time_matrix: List[List[int]]  # 旅行时间矩阵
    node_identifiers: List[int]         # 节点ID列表
    node_dict: Dict[int, Node]           # 节点ID到节点属性的映射


def load_instance() -> Instance:
    # 加载Excel工作簿
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    worksheet_nodes = workbook["节点属性信息"]
    worksheet_time = workbook["旅行时间矩阵"]

    # 读取节点属性，构建节点字典
    node_dict = {}
    for row in worksheet_nodes.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        node_id = int(row[0])
        node_dict[node_id] = Node(
            lower_bound=int(row[1]),
            upper_bound=int(row[2]),
            service_time=int(row[3]),
            demand=int(row[4]) if row[4] else 0
        )

    # 读取旅行时间矩阵的节点ID列表
    node_identifiers = [
        int(worksheet_time.cell(1, column).value)
        for column in range(2, worksheet_time.max_column + 1)
    ]

    # 读取旅行时间矩阵数据
    travel_time_matrix = [
        [
            int(worksheet_time.cell(row, column).value)
            for column in range(2, worksheet_time.max_column + 1)
        ]
        for row in range(2, worksheet_time.max_row + 1)
    ]

    return Instance(
        travel_time_matrix=travel_time_matrix,
        node_identifiers=node_identifiers,
        node_dict=node_dict
    )


def get_travel_time(from_node: int, to_node: int, instance: Instance) -> int:
    from_index = instance.node_identifiers.index(from_node)
    to_index = instance.node_identifiers.index(to_node)
    return instance.travel_time_matrix[from_index][to_index]


def route_evaluate(route: List[int], instance: Instance) -> dict:
    # 构建完整路线（首尾加上配送中心0）
    full_route = [0] + route + [0]

    current_time = 0           # 当前时间（累计旅行时间+服务时间）
    total_travel_time = 0     # 总运输时间
    total_penalty = 0         # 总惩罚值
    customer_details = []      # 每客户的详细数据

    # 遍历路线的每一段
    for prev_node, node in zip(full_route[:-1], full_route[1:], strict=True):
        # 获取该段旅行时间并累加到总运输时间
        leg_time = get_travel_time(prev_node, node, instance)
        total_travel_time += leg_time

        # 跳过配送中心（不需要计算时间窗惩罚）
        if node == 0:
            continue

        # 更新当前时间（到达该节点）
        current_time += leg_time
        service_start_time = int(current_time)

        # 获取该节点的时间窗信息
        node_info = instance.node_dict[node]
        lower_bound = node_info.lower_bound
        upper_bound = node_info.upper_bound

        # 计算时间窗违反量
        early_violation = max(lower_bound - service_start_time, 0)   # 早到违反时间
        late_violation = max(service_start_time - upper_bound, 0)     # 晚到违反时间

        # 计算惩罚值（二次惩罚）
        penalty_value = EARLY_PENALTY * early_violation ** 2 + LATE_PENALTY * late_violation ** 2

        # 累加总惩罚
        total_penalty += penalty_value

        # 记录该客户的详细数据
        customer_details.append({
            "node": node,
            "arrival": int(current_time),          # 到达时间
            "start": service_start_time,            # 开始服务时间
            "lb": lower_bound,                      # 时间窗下限
            "ub": upper_bound,                      # 时间窗上限
            "early": early_violation,               # 早到违反量
            "late": late_violation,                 # 晚到违反量
            "penalty": penalty_value               # 该客户惩罚值
        })

        # 加上服务时间
        current_time += node_info.service_time

    # 返回评估结果
    return {
        "route": full_route,
        "travel_time": total_travel_time,
        "time_window_penalty": total_penalty,
        "objective": total_travel_time + total_penalty,
        "per_customer": customer_details
    }


def insert_greedy(customer_order: List[int], instance: Instance) -> List[int]:
    route = []

    # 依次处理每个客户
    for customer in customer_order:
        if not route:
            # 第一个客户直接作为起点
            route = [customer]
            continue

        best_candidate = None
        best_key = None

        # 尝试所有可能的插入位置
        for position in range(len(route) + 1):
            candidate = route[:]
            candidate.insert(position, customer)
            evaluation = route_evaluate(candidate, instance)
            key = (evaluation["objective"], evaluation["travel_time"])

            if best_key is None or key < best_key:
                best_key = key
                best_candidate = candidate

        route = best_candidate

    return route


def nearest_neighbor(start_node: int, customer_list: List[int], instance: Instance) -> List[int]:
    remaining_customers = set(customer_list)
    current_node = start_node
    route = []

    # 循环直到所有客户都被访问
    while remaining_customers:
        # 选择距离最近的客户
        next_node = min(
            remaining_customers,
            key=lambda x: get_travel_time(current_node, x, instance)
        )
        route.append(next_node)
        remaining_customers.remove(next_node)
        current_node = next_node

    return route


def two_opt_swap(route: List[int], instance: Instance, max_iterations: int = 50) -> Tuple[List[int], dict]:
    best_route = route[:]
    best_value = route_evaluate(best_route, instance)["objective"]
    route_length = len(best_route)

    for _ in range(max_iterations):
        improved = False
        for i in range(route_length):
            for j in range(i + 1, route_length):
                # 反转i到j之间的节点
                candidate = best_route[:i] + list(reversed(best_route[i:j+1])) + best_route[j+1:]
                value = route_evaluate(candidate, instance)["objective"]
                if value + 1e-9 < best_value:
                    best_route = candidate
                    best_value = value
                    improved = True
        if not improved:
            break

    return best_route, route_evaluate(best_route, instance)


def relocate(route: List[int], instance: Instance) -> Tuple[List[int], dict]:
    best_route = route[:]
    best_value = route_evaluate(best_route, instance)["objective"]
    route_length = len(best_route)

    improved = True
    while improved:
        improved = False
        for i in range(route_length):
            for k in range(route_length + 1):
                if k == i or k == i + 1:
                    continue
                # 移除位置i的节点
                candidate = best_route[:i] + best_route[i+1:]
                insert_position = k if k <= i else k - 1
                candidate.insert(insert_position, best_route[i])
                value = route_evaluate(candidate, instance)["objective"]
                if value + 1e-9 < best_value:
                    best_route = candidate
                    best_value = value
                    improved = True

    return best_route, route_evaluate(best_route, instance)


def swap(route: List[int], instance: Instance) -> Tuple[List[int], dict]:
    best_route = route[:]
    best_value = route_evaluate(best_route, instance)["objective"]
    route_length = len(best_route)

    improved = True
    while improved:
        improved = False
        for i in range(route_length):
            for j in range(i + 1, route_length):
                candidate = best_route[:]
                candidate[i], candidate[j] = candidate[j], candidate[i]
                value = route_evaluate(candidate, instance)["objective"]
                if value + 1e-9 < best_value:
                    best_route = candidate
                    best_value = value
                    improved = True

    return best_route, route_evaluate(best_route, instance)


def combined_optimize(route: List[int], instance: Instance, max_iterations: int = 100) -> Tuple[List[int], dict]:
    best_route = route[:]
    best_value = route_evaluate(best_route, instance)["objective"]

    for iteration in range(max_iterations):
        improved = False

        # 2-opt交换
        current_route = best_route[:]
        current_value = best_value
        current_route, evaluation = two_opt_swap(current_route, instance, max_iterations=30)
        if evaluation["objective"] < current_value:
            best_route = current_route
            best_value = evaluation["objective"]
            improved = True
            continue

        # Relocate移动
        current_route, evaluation = relocate(current_route, instance)
        if evaluation["objective"] < current_value:
            best_route = current_route
            best_value = evaluation["objective"]
            improved = True
            continue

        # Swap移动
        current_route, evaluation = swap(current_route, instance)
        if evaluation["objective"] < current_value:
            best_route = current_route
            best_value = evaluation["objective"]
            improved = True
            continue

        if not improved:
            break

    return best_route, route_evaluate(best_route, instance)


def insert_violators_first(customer_list: List[int], instance: Instance) -> List[int]:
    # 找出最容易晚到的客户（时间窗最紧的）
    customer_scores = []
    for customer in customer_list:
        lower_bound = instance.node_dict[customer].lower_bound
        upper_bound = instance.node_dict[customer].upper_bound
        # 紧迫度 = upper_bound（越早的时间窗越紧迫）
        score = upper_bound
        customer_scores.append((customer, score))

    # 按紧迫度排序
    customer_scores.sort(key=lambda x: x[1])

    route = []
    for customer, _ in customer_scores:
        if not route:
            route = [customer]
            continue

        # 找最优插入位置（考虑对已有路线的影响）
        best = None
        for position in range(len(route) + 1):
            candidate = route[:]
            candidate.insert(position, customer)
            evaluation = route_evaluate(candidate, instance)
            key = (evaluation["objective"], evaluation["travel_time"])
            if best is None or key < best[0]:
                best = (key, candidate)
        route = best[1]

    return route


def solve_by_time_center(instance: Instance) -> dict:
    order = sorted(
        CUSTOMERS,
        key=lambda c: (instance.node_dict[c].lower_bound + instance.node_dict[c].upper_bound) / 2
    )
    route = insert_greedy(order, instance)
    route, evaluation = combined_optimize(route, instance)
    return route, evaluation


def solve_by_earliest_upper_bound(instance: Instance) -> dict:
    order = sorted(
        CUSTOMERS,
        key=lambda c: (instance.node_dict[c].upper_bound, instance.node_dict[c].lower_bound)
    )
    route = insert_greedy(order, instance)
    route, evaluation = combined_optimize(route, instance)
    return route, evaluation


def solve_by_violator_priority(instance: Instance) -> dict:
    route = insert_violators_first(CUSTOMERS, instance)
    route, evaluation = combined_optimize(route, instance)
    return route, evaluation


def solve_by_nearest_neighbor(instance: Instance) -> dict:
    best_route = None
    best_evaluation = None

    # 尝试不同的起始点
    for start_customer in CUSTOMERS:
        route = nearest_neighbor(start_customer, CUSTOMERS, instance)
        route, evaluation = combined_optimize(route, instance)
        if best_evaluation is None or evaluation["objective"] < best_evaluation["objective"]:
            best_route = route
            best_evaluation = evaluation

    return best_route, best_evaluation


def solve_by_segment_decomposition(instance: Instance) -> dict:
    # 按upper_bound排序划分5段
    sorted_customers = sorted(CUSTOMERS, key=lambda c: instance.node_dict[c].upper_bound)
    band_size = 10
    bands = [sorted_customers[i:i+band_size] for i in range(0, len(sorted_customers), band_size)]

    print(f"  划分{len(bands)}个时段")

    # 段内求解
    band_routes = []
    for band_index, band in enumerate(bands):
        order = sorted(band, key=lambda c: instance.node_dict[c].upper_bound)
        route = insert_greedy(order, instance)
        route, _ = combined_optimize(route, instance)
        band_routes.append(route)

    # 合并：尝试所有排列组合
    from itertools import permutations
    best_full_route = []
    best_full_evaluation = None

    num_bands = len(band_routes)
    for perm in permutations(range(num_bands)):
        for orientations in product([0, 1], repeat=num_bands):
            full_route = []
            for band_index, orientation in zip(perm, orientations):
                segment = band_routes[band_index][:]
                if orientation == 1:
                    segment = list(reversed(segment))
                full_route.extend(segment)

            evaluation = route_evaluate(full_route, instance)
            if best_full_evaluation is None or evaluation["objective"] < best_full_evaluation["objective"]:
                best_full_evaluation = evaluation
                best_full_route = full_route

    # 全局优化
    best_full_route, best_full_evaluation = combined_optimize(best_full_route, instance)

    return best_full_route, best_full_evaluation


def solve_q3(instance: Instance) -> dict:
    print("=" * 70)
    print("问题3：50客户大规模带时间窗调度")
    print("策略：多策略初始解 + 组合局部搜索")
    print("=" * 70)

    all_results = {}

    # 方法1：时间窗中心排序
    print("\n--- 方法1: 时间窗中心排序 ---")
    route, evaluation = solve_by_time_center(instance)
    all_results["time_center"] = evaluation
    print(f"  obj={evaluation['objective']}, travel={evaluation['travel_time']}, pen={evaluation['time_window_penalty']}")

    # 方法2：最早upper_bound排序
    print("\n--- 方法2: 最早upper_bound排序 ---")
    route, evaluation = solve_by_earliest_upper_bound(instance)
    all_results["earliest_upper_bound"] = evaluation
    print(f"  obj={evaluation['objective']}, travel={evaluation['travel_time']}, pen={evaluation['time_window_penalty']}")

    # 方法3：紧迫客户优先
    print("\n--- 方法3: 紧迫客户优先 ---")
    route, evaluation = solve_by_violator_priority(instance)
    all_results["violator_priority"] = evaluation
    print(f"  obj={evaluation['objective']}, travel={evaluation['travel_time']}, pen={evaluation['time_window_penalty']}")

    # 方法4：最近邻
    print("\n--- 方法4: 最近邻构造 ---")
    route, evaluation = solve_by_nearest_neighbor(instance)
    all_results["nearest_neighbor"] = evaluation
    print(f"  obj={evaluation['objective']}, travel={evaluation['travel_time']}, pen={evaluation['time_window_penalty']}")

    # 方法5：时间段分解
    print("\n--- 方法5: 时间段分解 ---")
    route, evaluation = solve_by_segment_decomposition(instance)
    all_results["segment_decomposition"] = evaluation
    print(f"  obj={evaluation['objective']}, travel={evaluation['travel_time']}, pen={evaluation['time_window_penalty']}")

    # 取最优
    best = min(all_results.values(), key=lambda x: x["objective"])

    print("\n" + "=" * 70)
    print(f"最优解: obj={best['objective']}, travel={best['travel_time']}, pen={best['time_window_penalty']}")
    print(f"路线: {best['route']}")

    return best


def main():
    # 加载实例数据
    instance = load_instance()
    print(f"\n数据加载: 客户1-50, 早={EARLY_PENALTY}, 晚={LATE_PENALTY}")

    # 求解
    result = solve_q3(instance)

    # 保存结果
    payload = {
        "problem": 3,
        "scope": {
            "depot": 0,
            "customer_ids": CUSTOMERS,
            "customer_count": len(CUSTOMERS),
            "consider_time_window_penalty": True,
            "consider_capacity": False
        },
        "strategy": {
            "name": "多策略初始解 + 组合局部搜索",
            "description": "5种初始解构造方法 + 2-opt + relocate + swap多轮迭代",
        },
        "final_answer": result,
    }

    output_path = Path(r"output_json/q3_result.json")
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n结果已保存: {output_path}")


if __name__ == "__main__":
    main()
